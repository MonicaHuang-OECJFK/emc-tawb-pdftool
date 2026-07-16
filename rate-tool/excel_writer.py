"""
excel_writer.py (EMC 版)
────────────────────────
把從 EMC PDF 萃取的 2SD/4SD/4SH 寫入 cheatsheet 的 rate sheet。

跟 COSCO 那份不同：
  - 不需要 Mapping tab，因為 PDF 萃取出來的 POL/POD 字串
    跟 cheatsheet 裡的 POL/POD 是一樣的，直接比對即可。
  - 欄位結構：POD 往右 +1 = 20'，+2 = 40'，+3 = 40'HC
    （不是 +2 起跳，跟 COSCO 那份不一樣，寫死前務必用這份 PDF 重新驗證）
  - 資料格沒有公式，全部直接覆寫。

流程：
1. 掃描整張 sheet 找 "POL" / "POD" header cell → 拿到 col + header_row
2. POD col +1 = 20' 候選欄，往上掃該欄有沒有任何 row 是 "20'" 來驗證
   （避免欄位結構改動後寫錯地方而不自知）
3. 確認後 +2 = 40'，+3 = 40'HC
4. 從 header_row+1 開始掃資料列，直到 POL/POD 都是空白才停
5. 用 (POL, POD) 大寫字串直接比對 PDF 萃取的 rates，寫入 20'/40'/40'HC

依賴：pip install openpyxl
"""

import openpyxl


def _find_header(ws, keyword):
    """掃描整張 sheet，找第一個 value == keyword 的 cell。回傳 (row, col)。"""
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip().upper() == keyword.upper():
                return cell.row, cell.column
    raise ValueError(f"找不到 header「{keyword}」")


def _find_rate_cols(ws, pod_col, header_row):
    """
    從 POD 欄往右 +1 找 20' 欄，往上掃這欄驗證有沒有 "20'" header。
    確認後 +2 = 40'，+3 = 40'HC。

    回傳 (col_20, col_40, col_40hc)。
    """
    col_20_candidate = pod_col + 1

    confirmed = False
    for r in range(1, header_row + 1):
        val = str(ws.cell(r, col_20_candidate).value or "").strip()
        if val == "20'":
            confirmed = True
            break

    if not confirmed:
        raise ValueError(
            f"POD 右邊 +1 格（col {col_20_candidate}）往上找不到 \"20'\" header，"
            f"請確認 cheatsheet 欄位結構是否改動"
        )

    col_20   = col_20_candidate
    col_40   = col_20 + 1
    col_40hc = col_20 + 2

    return col_20, col_40, col_40hc


def _build_rate_lookup(all_rates):
    """
    把 PDF 萃取的 all_rates（list of dict，含 pol/pod/2sd/4sd/4sh）
    轉成 {(POL_upper, POD_upper): (2sd, 4sd, 4sh)} 字典。
    """
    lookup = {}
    for row in all_rates:
        pol = row["pol"].strip().upper()
        pod = row["pod"].strip().upper()
        lookup[(pol, pod)] = (row["2sd"], row["4sd"], row["4sh"])
    return lookup


def update_rates(excel_path, all_rates, output_path=None, sheet_name="rate sheet"):
    """
    主函式：把 all_rates 寫入 cheatsheet。

    參數：
      excel_path  : cheatsheet 路徑
      all_rates   : emc_extract.extract() 產出的 list of dict
                    每筆含 pol / pod / 2sd / 4sd / 4sh
      output_path : 輸出路徑，None 則直接覆蓋原檔
      sheet_name  : 要寫入的分頁名稱

    回傳：
      (updated_count, skipped_rows)
      updated_count : 成功寫入的列數
      skipped_rows  : 找不到對應 rate 的 (row_num, pol, pod) list
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]

    rate_lookup = _build_rate_lookup(all_rates)

    pol_header_row, pol_col = _find_header(ws, "POL")
    pod_header_row, pod_col = _find_header(ws, "POD")
    col_20, col_40, col_40hc = _find_rate_cols(ws, pod_col, pod_header_row)

    data_start_row = max(pol_header_row, pod_header_row) + 1

    updated_count = 0
    skipped_rows = []

    row_num = data_start_row
    while row_num <= ws.max_row:
        pol = str(ws.cell(row_num, pol_col).value or "").strip().upper()
        pod = str(ws.cell(row_num, pod_col).value or "").strip().upper()

        if not pol and not pod:
            row_num += 1
            continue

        key = (pol, pod)
        if key not in rate_lookup:
            skipped_rows.append((row_num, pol, pod))
            row_num += 1
            continue

        rate_2sd, rate_4sd, rate_4sh = rate_lookup[key]
        ws.cell(row_num, col_20).value   = rate_2sd
        ws.cell(row_num, col_40).value   = rate_4sd
        ws.cell(row_num, col_40hc).value = rate_4sh

        updated_count += 1
        row_num += 1

    out = output_path or excel_path
    wb.save(out)
    return updated_count, skipped_rows
