"""
emc_to_excel.py
───────────────
從 Evergreen (EMC) PDF 萃取費率並輸出成 Excel。

用法：python emc_to_excel.py <path_to_pdf>
輸出：emc_extract.xlsx（執行目錄）

依賴：pip install pdfplumber openpyxl
"""

import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from emc_extract import extract


HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

HEADERS = ["POL", "POD", "2SD (USD)", "4SD (USD)", "4SH (USD)"]
KEYS    = ["pol", "pod", "2sd", "4sd", "4sh"]


def write_sheet(ws, rows):
    ws.title = "Rates"

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 20

    for r_idx, row in enumerate(rows, 2):
        for c_idx, key in enumerate(KEYS, 1):
            cell = ws.cell(r_idx, c_idx, row[key])
            cell.font      = DATA_FONT
            cell.alignment = LEFT if c_idx <= 2 else CENTER

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14


def main():
    if len(sys.argv) < 2:
        print("Usage: python emc_to_excel.py <path_to_pdf>")
        sys.exit(1)

    pdf_path = sys.argv[1]
    rows = extract(pdf_path)

    wb = Workbook()
    write_sheet(wb.active, rows)

    out_path = os.path.join(os.getcwd(), "emc_extract.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}  ({len(rows)} rows)")


if __name__ == "__main__":
    main()
